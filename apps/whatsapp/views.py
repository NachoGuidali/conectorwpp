import json
import logging
import re

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import ListView
from django.views.decorators.csrf import csrf_exempt

from apps.users.models import User
from .models import Conversacion, Mensaje, PlantillaHSM, ConfiguracionWhatsApp
from .tasks import process_incoming_message, send_whatsapp_message_task
from .webhook import parse_incoming_webhook, verify_webhook_token

logger = logging.getLogger('apps.whatsapp')


class SupervisorRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_supervisor:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Solo supervisores y administradores.')
        return super().dispatch(request, *args, **kwargs)


def _get_convs_qs(user, include_archived=False):
    from django.db.models import Q
    qs = Conversacion.objects.select_related('agente').order_by('-ultimo_mensaje_at', '-pk')
    if not include_archived:
        qs = qs.filter(archivada=False)
    if not user.can_see_all:
        qs = qs.filter(agente=user)
    return qs


@method_decorator(csrf_exempt, name='dispatch')
class WebhookView(View):
    def get(self, request):
        return HttpResponse('OK', status=200)

    def post(self, request):
        token = request.headers.get('apikey', '')
        configured_token = ConfiguracionWhatsApp.get_setting('webhook_token')
        # Accept both the configured token AND the Evolution API instance token
        evo_api_key = ConfiguracionWhatsApp.get_setting('evolution_api_key')
        if not verify_webhook_token(token, configured_token) and token != evo_api_key:
            return HttpResponse('Forbidden', status=403)
        try:
            payload = json.loads(request.body)
            event = payload.get('event', '')
            # Cache QR code delivered by webhook
            if event in ('QRCODE_UPDATED', 'qrcode.updated'):
                from django.core.cache import cache
                qr_data = payload.get('data', {})
                qr_obj = qr_data.get('qrcode', qr_data)
                qr_code = qr_obj.get('code', '')
                qr_b64 = qr_obj.get('base64', '')
                if qr_b64 and ',' in qr_b64:
                    qr_b64 = qr_b64.split(',', 1)[1]
                if qr_code:
                    cache.set('whatsapp_qr_code', qr_b64, timeout=55)
                    cache.set('whatsapp_qr_text', qr_code, timeout=55)
                    logger.info('QR code cached from webhook (code len=%d)', len(qr_code))
                return HttpResponse('OK', status=200)
            messages_data = parse_incoming_webhook(payload)
            for msg_data in messages_data:
                process_incoming_message.delay(msg_data)
        except Exception as e:
            logger.exception('Webhook error: %s', e)
        return HttpResponse('OK', status=200)


class InboxView(LoginRequiredMixin, View):
    template_name = 'whatsapp/inbox.html'

    def get(self, request):
        from django.db.models import Q
        qs = _get_convs_qs(request.user)

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(nombre_contacto__icontains=q) | Q(telefono__icontains=q))

        sin_agente = request.GET.get('sin_agente', '').strip()
        if sin_agente:
            qs = qs.filter(agente__isnull=True)

        solo_no_leidos = request.GET.get('no_leidos', '').strip()
        if solo_no_leidos:
            qs = qs.filter(mensajes_no_leidos__gt=0)

        sin_contacto = request.GET.get('sin_contacto', '').strip()
        if sin_contacto:
            qs = qs.filter(contacto__isnull=True)

        archivadas = request.GET.get('archivadas', '').strip()
        if archivadas:
            # Mostrar archivadas en lugar de activas
            qs = Conversacion.objects.filter(archivada=True)
            if not request.user.can_see_all:
                qs = qs.filter(agente=request.user)
            if q:
                qs = qs.filter(Q(nombre_contacto__icontains=q) | Q(telefono__icontains=q))
            qs = qs.order_by('-ultimo_mensaje_at')

        conversaciones = list(qs[:100])
        unread_total = _get_convs_qs(request.user).filter(mensajes_no_leidos__gt=0).count()

        # Conversación seleccionada
        selected_conv = None
        mensajes = []
        plantillas = []
        agents = None
        last_msg_id = 0

        contacto_campos = []
        conv_pk = request.GET.get('conv', '').strip()
        if conv_pk:
            try:
                # Si viene del filtro archivadas, buscar también en archivadas
                if archivadas:
                    conv_qs = Conversacion.objects.filter(archivada=True)
                    if not request.user.can_see_all:
                        conv_qs = conv_qs.filter(agente=request.user)
                else:
                    conv_qs = _get_convs_qs(request.user)
                selected_conv = (
                    conv_qs
                    .select_related('contacto')
                    .get(pk=int(conv_pk))
                )
                Conversacion.objects.filter(pk=selected_conv.pk).update(mensajes_no_leidos=0)
                msgs_qs = selected_conv.mensajes.order_by('timestamp')
                total = msgs_qs.count()
                mensajes = list(msgs_qs[max(0, total - 60):])
                plantillas = PlantillaHSM.objects.filter(activa=True)
                if request.user.can_see_all:
                    agents = User.objects.filter(is_active=True).order_by('first_name', 'username')
                last_msg = selected_conv.mensajes.order_by('timestamp').last()
                last_msg_id = last_msg.pk if last_msg else 0
                if selected_conv.contacto:
                    contacto_campos = selected_conv.contacto.get_campos_con_valores()
            except (Conversacion.DoesNotExist, ValueError):
                selected_conv = None

        return render(request, self.template_name, {
            'conversaciones': conversaciones,
            'unread_total': unread_total,
            'q': q,
            'sin_agente': sin_agente,
            'solo_no_leidos': solo_no_leidos,
            'sin_contacto': sin_contacto,
            'archivadas': archivadas,
            'selected_conv': selected_conv,
            'mensajes': mensajes,
            'plantillas': plantillas,
            'agents': agents,
            'last_msg_id': last_msg_id,
            'contacto_campos': contacto_campos,
        })

    def post(self, request):
        conv_pk = request.POST.get('conv_pk', '').strip()
        if not conv_pk:
            return redirect('whatsapp:inbox')

        conv = get_object_or_404(_get_convs_qs(request.user, include_archived=True), pk=conv_pk)
        action = request.POST.get('action', '')

        if action == 'send_text':
            body = request.POST.get('body', '').strip()
            if body:
                msg = Mensaje.objects.create(
                    conversacion=conv, direccion=Mensaje.DIR_SALIENTE,
                    tipo=Mensaje.TIPO_TEXTO, contenido=body,
                    status=Mensaje.STATUS_PENDIENTE,
                    enviado_por=request.user, timestamp=timezone.now(),
                )
                send_whatsapp_message_task.delay(msg.pk)
                Conversacion.objects.filter(pk=conv.pk).update(ultimo_mensaje_at=timezone.now())

        elif action == 'send_template':
            plantilla_id = request.POST.get('plantilla_id')
            if plantilla_id:
                plantilla = get_object_or_404(PlantillaHSM, pk=plantilla_id)
                vals = [request.POST.get(f'var_{i+1}', '') for i in range(len(plantilla.variables or []))]
                text = plantilla.preview(vals if any(vals) else None)
                try:
                    from .sender import send_text_message
                    result = send_text_message(conv.telefono, text)
                    Mensaje.objects.create(
                        conversacion=conv, direccion=Mensaje.DIR_SALIENTE,
                        tipo=Mensaje.TIPO_PLANTILLA, contenido=text,
                        whatsapp_message_id=result.get('id', ''),
                        status=Mensaje.STATUS_ENVIADO,
                        enviado_por=request.user, timestamp=timezone.now(),
                    )
                    Conversacion.objects.filter(pk=conv.pk).update(ultimo_mensaje_at=timezone.now())
                    messages.success(request, 'Plantilla enviada.')
                except Exception as e:
                    messages.error(request, f'Error: {e}')

        params = f'conv={conv.pk}'
        if request.POST.get('_q'):
            params += f'&q={request.POST.get("_q")}'
        if request.POST.get('_archivadas'):
            params += '&archivadas=1'
        if request.POST.get('_sin_contacto'):
            params += '&sin_contacto=1'
        from django.urls import reverse
        return redirect(f"{reverse('whatsapp:inbox')}?{params}")


class ConversacionMessagesAPIView(LoginRequiredMixin, View):
    def get(self, request, pk):
        conv = get_object_or_404(_get_convs_qs(request.user, include_archived=True), pk=pk)
        since_id = int(request.GET.get('since_id', 0))
        nuevos = conv.mensajes.filter(pk__gt=since_id).order_by('timestamp')
        if nuevos.exists():
            Conversacion.objects.filter(pk=pk).update(mensajes_no_leidos=0)
        return JsonResponse({'mensajes': [{
            'id': m.pk, 'direccion': m.direccion, 'tipo': m.tipo,
            'contenido': m.contenido, 'media_url': m.media_url,
            'media_filename': m.media_filename, 'media_mime': m.media_mime,
            'status': m.status, 'borrado': m.borrado,
            'timestamp': m.timestamp.strftime('%d/%m %H:%M'),
            'enviado_por': m.enviado_por.get_full_name() if m.enviado_por else '',
        } for m in nuevos]})


class DashboardSupervisorView(LoginRequiredMixin, View):
    template_name = 'whatsapp/dashboard_supervisor.html'

    def get(self, request):
        if not request.user.can_see_all:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()

        from django.contrib.auth import get_user_model
        from django.db.models import Count, Q as Qm
        User = get_user_model()

        agentes = (
            User.objects.filter(rol=User.ROL_AGENTE)
            .annotate(
                total=Count('conversaciones', filter=Qm(conversaciones__archivada=False)),
                bot=Count('conversaciones', filter=Qm(conversaciones__archivada=False, conversaciones__bot_n8n_activo=True)),
                pendiente=Count('conversaciones', filter=Qm(conversaciones__archivada=False, conversaciones__estado='pendiente')),
                abierta=Count('conversaciones', filter=Qm(conversaciones__archivada=False, conversaciones__estado='abierta', conversaciones__bot_n8n_activo=False)),
            )
            .order_by('-en_turno', 'username')
        )

        sin_asignar = Conversacion.objects.filter(
            agente__isnull=True, archivada=False
        ).order_by('-ultimo_mensaje_at')[:20]

        sin_contacto = Conversacion.objects.filter(
            contacto__isnull=True, archivada=False
        ).select_related('agente').order_by('-ultimo_mensaje_at')[:50]

        # Detalle de convs por agente (para el panel expandible)
        agente_pk = request.GET.get('agente')
        convs_agente = []
        agente_sel = None
        if agente_pk:
            try:
                agente_sel = User.objects.get(pk=agente_pk, rol=User.ROL_AGENTE)
                convs_agente = Conversacion.objects.filter(
                    agente=agente_sel, archivada=False
                ).order_by('-ultimo_mensaje_at').select_related('contacto')[:50]
            except User.DoesNotExist:
                pass

        return render(request, self.template_name, {
            'agentes': agentes,
            'sin_asignar': sin_asignar,
            'sin_contacto': sin_contacto,
            'agente_sel': agente_sel,
            'convs_agente': convs_agente,
            'todos_agentes': User.objects.filter(rol=User.ROL_AGENTE, is_active=True).order_by('username'),
        })

    def post(self, request):
        """Reasignar todas las conversaciones de un agente a otro, o togglear su disponibilidad para la cola."""
        if not request.user.can_see_all:
            return JsonResponse({'ok': False, 'error': 'Sin permisos'}, status=403)

        from django.contrib.auth import get_user_model
        User = get_user_model()

        toggle_pk = request.POST.get('toggle_recibe_pk')
        if toggle_pk:
            agente = get_object_or_404(User, pk=toggle_pk, rol=User.ROL_AGENTE)
            agente.recibe_asignaciones = not agente.recibe_asignaciones
            agente.save(update_fields=['recibe_asignaciones'])
            return redirect(request.POST.get('next') or request.path)

        desde_pk = request.POST.get('desde_agente')
        hacia_pk = request.POST.get('hacia_agente') or None

        convs = Conversacion.objects.filter(agente_id=desde_pk, archivada=False)
        if hacia_pk:
            convs.update(agente_id=hacia_pk)
            msg = f'{convs.count()} conversaciones reasignadas.'
        else:
            # Redistribuir automáticamente
            from apps.whatsapp.tasks import auto_asignar_agente
            pks = list(convs.values_list('pk', flat=True))
            convs.update(agente=None)
            for conv in Conversacion.objects.filter(pk__in=pks):
                auto_asignar_agente(conv)
            msg = f'{len(pks)} conversaciones redistribuidas automáticamente.'

        from django.contrib import messages as msgs
        msgs.success(request, msg)
        return redirect(f"{request.path}?agente={desde_pk}")


class ConversacionesExportarView(LoginRequiredMixin, View):
    def get(self, request):
        if not request.user.can_see_all:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        from django.utils.dateparse import parse_date
        fecha_desde = request.GET.get('desde', '').strip()
        fecha_hasta = request.GET.get('hasta', '').strip()
        origen = request.GET.get('origen', '').strip()

        qs = (
            Conversacion.objects.all()
            .select_related('agente', 'contacto')
            .order_by('-ultimo_mensaje_at')
        )

        if fecha_desde:
            d = parse_date(fecha_desde)
            if d:
                qs = qs.filter(created_at__date__gte=d)
        if fecha_hasta:
            d = parse_date(fecha_hasta)
            if d:
                qs = qs.filter(created_at__date__lte=d)
        if origen in ('entrante', 'saliente'):
            qs = qs.filter(origen_conversacion=origen)
        archivada_filter = request.GET.get('archivada', '').strip()
        if archivada_filter == 'si':
            qs = qs.filter(archivada=True)
        elif archivada_filter == 'no':
            qs = qs.filter(archivada=False)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Conversaciones'
        headers = ['Nombre', 'Teléfono', 'Agente', 'Email agente', 'Estado', 'Archivada', 'Origen', 'Último mensaje', 'Creado']
        ws.append(headers)

        estado_labels = dict(Conversacion.ESTADO_CHOICES)
        for conv in qs:
            agente_nombre = ''
            agente_email = ''
            if conv.agente:
                agente_nombre = f"{conv.agente.first_name} {conv.agente.last_name}".strip() or conv.agente.username
                agente_email = conv.agente.email or ''
            nombre = ''
            if conv.contacto:
                nombre = conv.contacto.nombre
            nombre = nombre or conv.nombre_contacto or conv.telefono
            ws.append([
                nombre,
                conv.telefono,
                agente_nombre,
                agente_email,
                estado_labels.get(conv.estado, conv.estado),
                'Sí' if conv.archivada else 'No',
                dict(Conversacion.ORIGEN_CHOICES).get(conv.origen_conversacion, conv.origen_conversacion),
                conv.ultimo_mensaje_at.strftime('%d/%m/%Y %H:%M') if conv.ultimo_mensaje_at else '',
                conv.created_at.strftime('%d/%m/%Y %H:%M') if conv.created_at else '',
            ])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 24

        filename = 'conversaciones'
        if fecha_desde or fecha_hasta:
            filename += f'_{fecha_desde or "inicio"}_{fecha_hasta or "hoy"}'
        filename += '.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class SinContactoExportarView(LoginRequiredMixin, View):
    def get(self, request):
        if not request.user.can_see_all:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        qs = (
            Conversacion.objects.filter(contacto__isnull=True, archivada=False)
            .select_related('agente')
            .order_by('-ultimo_mensaje_at')
        )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sin contacto'
        headers = ['Nombre / Teléfono', 'Teléfono', 'Agente', 'Estado', 'Último mensaje', 'Creado']
        ws.append(headers)

        estado_labels = dict(Conversacion.ESTADO_CHOICES)
        for conv in qs:
            agente_nombre = ''
            if conv.agente:
                agente_nombre = f"{conv.agente.first_name} {conv.agente.last_name}".strip() or conv.agente.username
            ws.append([
                conv.nombre_contacto or conv.telefono,
                conv.telefono,
                agente_nombre,
                estado_labels.get(conv.estado, conv.estado),
                conv.ultimo_mensaje_at.strftime('%d/%m/%Y %H:%M') if conv.ultimo_mensaje_at else '',
                conv.created_at.strftime('%d/%m/%Y %H:%M') if conv.created_at else '',
            ])

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 24

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="conversaciones_sin_contacto.xlsx"'
        wb.save(response)
        return response


class ReporteConversacionesView(LoginRequiredMixin, View):
    template_name = 'whatsapp/reporte_conversaciones.html'

    def _rango(self, request):
        import datetime
        hoy = timezone.localdate()
        desde_str = request.GET.get('desde', '')
        hasta_str = request.GET.get('hasta', '')
        try:
            desde = datetime.date.fromisoformat(desde_str) if desde_str else hoy - datetime.timedelta(days=29)
        except ValueError:
            desde = hoy - datetime.timedelta(days=29)
        try:
            hasta = datetime.date.fromisoformat(hasta_str) if hasta_str else hoy
        except ValueError:
            hasta = hoy
        return desde, hasta

    def _datos(self, desde, hasta):
        from django.db.models import Count
        from django.db.models.functions import TruncDate

        qs = Conversacion.objects.filter(created_at__date__gte=desde, created_at__date__lte=hasta)

        por_dia = list(
            qs.annotate(dia=TruncDate('created_at'))
            .values('dia').annotate(total=Count('pk')).order_by('dia')
        )
        por_agente = list(
            qs.values('agente__username', 'agente__first_name', 'agente__last_name')
            .annotate(total=Count('pk')).order_by('-total')
        )
        por_estado = list(
            qs.values('estado').annotate(total=Count('pk')).order_by('-total')
        )
        por_origen = list(
            qs.values('origen_conversacion').annotate(total=Count('pk')).order_by('origen_conversacion')
        )
        estado_labels = dict(Conversacion.ESTADO_CHOICES)
        origen_labels = dict(Conversacion.ORIGEN_CHOICES)
        for row in por_estado:
            row['estado_label'] = estado_labels.get(row['estado'], row['estado'])
        for row in por_origen:
            row['origen_label'] = origen_labels.get(row['origen_conversacion'], row['origen_conversacion'])
        for row in por_agente:
            nombre = f"{row['agente__first_name']} {row['agente__last_name']}".strip()
            row['agente_nombre'] = nombre or row['agente__username'] or 'Sin agente'

        return {
            'total': qs.count(),
            'por_dia': por_dia,
            'por_agente': por_agente,
            'por_estado': por_estado,
            'por_origen': por_origen,
        }

    def get(self, request):
        if not request.user.can_see_all:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()

        desde, hasta = self._rango(request)
        datos = self._datos(desde, hasta)

        return render(request, self.template_name, {
            'desde': desde,
            'hasta': hasta,
            **datos,
        })


class ReporteConversacionesExportarView(ReporteConversacionesView):
    def get(self, request):
        if not request.user.can_see_all:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()

        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        desde, hasta = self._rango(request)
        datos = self._datos(desde, hasta)

        wb = Workbook()

        ws = wb.active
        ws.title = 'Por día'
        ws.append(['Fecha', 'Cantidad de conversaciones'])
        for row in datos['por_dia']:
            ws.append([row['dia'].strftime('%d/%m/%Y'), row['total']])
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 26

        ws2 = wb.create_sheet('Por agente')
        ws2.append(['Agente', 'Cantidad de conversaciones'])
        for row in datos['por_agente']:
            ws2.append([row['agente_nombre'], row['total']])
        ws2.column_dimensions['A'].width = 26
        ws2.column_dimensions['B'].width = 26

        ws3 = wb.create_sheet('Por estado')
        ws3.append(['Estado', 'Cantidad de conversaciones'])
        for row in datos['por_estado']:
            ws3.append([row['estado_label'], row['total']])
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 26

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre = f'conversaciones_{desde.isoformat()}_a_{hasta.isoformat()}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        wb.save(response)
        return response


class DashboardAgenteView(LoginRequiredMixin, View):
    template_name = 'whatsapp/dashboard_agente.html'

    def get(self, request):
        from django.db.models import Q as Qm
        convs = Conversacion.objects.filter(
            agente=request.user, archivada=False
        ).order_by('-ultimo_mensaje_at').select_related('contacto')

        bot_activo = convs.filter(bot_n8n_activo=True)
        pendientes = convs.filter(estado=Conversacion.ESTADO_PENDIENTE)
        abiertas = convs.filter(
            estado=Conversacion.ESTADO_ABIERTA, bot_n8n_activo=False
        )
        cerradas_hoy = Conversacion.objects.filter(
            agente=request.user,
            estado=Conversacion.ESTADO_CERRADA,
            ultimo_mensaje_at__date=timezone.now().date(),
        ).count()

        return render(request, self.template_name, {
            'bot_activo': bot_activo,
            'pendientes': pendientes,
            'abiertas': abiertas,
            'cerradas_hoy': cerradas_hoy,
            'total': convs.count(),
        })


class InboxUpdatesAPIView(LoginRequiredMixin, View):
    def get(self, request):
        qs = _get_convs_qs(request.user).filter(mensajes_no_leidos__gt=0)
        return JsonResponse({
            'unread_total': qs.count(),
            'conv_ids': list(qs.values_list('id', flat=True)),
        })


class InboxSSEView(LoginRequiredMixin, View):
    """
    Fast-polling SSE: responde en <100ms con los eventos disponibles y cierra.
    El browser reconecta cada 1.5s via EventSource retry.
    Cada worker solo está ocupado <100ms por request, nunca bloqueado.
    """

    def get(self, request):
        conv_pk = request.GET.get('conv_pk') or None
        last_msg_id = int(request.headers.get('Last-Event-ID') or
                          request.GET.get('last_msg_id') or 0)
        last_conv_ts = request.GET.get('last_conv_ts') or '0'

        events = []

        # ── Nuevos mensajes en la conversación activa ──────────────────
        if conv_pk:
            try:
                nuevos = list(
                    Mensaje.objects.filter(
                        conversacion_id=conv_pk,
                        pk__gt=last_msg_id,
                    ).order_by('pk').select_related('enviado_por')[:20]
                )
                for m in nuevos:
                    last_msg_id = m.pk
                    data = json.dumps({
                        'id': m.pk, 'tipo': m.tipo,
                        'contenido': m.contenido,
                        'direccion': m.direccion,
                        'media_url': m.media_url,
                        'media_filename': m.media_filename,
                        'media_mime': m.media_mime,
                        'timestamp': m.timestamp.strftime('%H:%M'),
                        'enviado_por': m.enviado_por.get_full_name() if m.enviado_por else '',
                    })
                    events.append(f'id: {m.pk}\nevent: message\ndata: {data}\n\n')
            except Exception:
                pass

        # ── Lista de conversaciones ────────────────────────────────────
        try:
            convs = list(
                _get_convs_qs(request.user)
                .order_by('-ultimo_mensaje_at')[:30]
                .values('pk', 'nombre_contacto', 'telefono',
                        'mensajes_no_leidos', 'ultimo_mensaje_at',
                        'archivada', 'estado')
            )
            conv_hash = str(hash(str([
                (c['pk'], c['mensajes_no_leidos'], str(c['ultimo_mensaje_at']), c['estado'])
                for c in convs
            ])))
            if conv_hash != last_conv_ts:
                for c in convs:
                    c['ultimo_mensaje_at'] = (
                        c['ultimo_mensaje_at'].strftime('%d/%m %H:%M')
                        if c['ultimo_mensaje_at'] else ''
                    )
                events.append(f'event: conv_list\ndata: {json.dumps({"convs": convs, "hash": conv_hash})}\n\n')
        except Exception:
            conv_hash = last_conv_ts

        # retry: 1500 → browser reconecta cada 1.5s
        body = 'retry: 1500\n\n' + ''.join(events)
        if not events:
            body += ': poll\n\n'

        response = HttpResponse(body, content_type='text/event-stream; charset=utf-8')
        response['Cache-Control'] = 'no-cache'
        response['X-Accel-Buffering'] = 'no'
        return response


class NuevaConversacionView(LoginRequiredMixin, View):
    def post(self, request):
        telefono = request.POST.get('telefono', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        contacto_id = request.POST.get('contacto_id', '').strip()

        if not telefono:
            messages.error(request, 'Ingresá un número de teléfono.')
            return redirect('whatsapp:inbox')
        if not telefono.startswith('+'):
            telefono = '+' + telefono

        contacto = None
        try:
            from apps.contacts.models import Contacto
            if contacto_id:
                contacto = Contacto.objects.get(pk=int(contacto_id))
                telefono = contacto.telefono
                nombre = contacto.nombre
            else:
                contacto, _ = Contacto.objects.get_or_create(
                    telefono=telefono,
                    defaults={'nombre': nombre or telefono},
                )
                if contacto.nombre == contacto.telefono and nombre:
                    contacto.nombre = nombre
                    contacto.save(update_fields=['nombre'])
                nombre = nombre or contacto.nombre
        except Exception:
            pass

        conv, _ = Conversacion.objects.get_or_create(
            telefono=telefono,
            defaults={
                'nombre_contacto': nombre or telefono,
                'agente': request.user,
                'contacto': contacto,
                'origen_conversacion': Conversacion.ORIGEN_SALIENTE,
            },
        )
        update_fields = []
        if not conv.contacto and contacto:
            conv.contacto = contacto
            update_fields.append('contacto')
        if conv.archivada:
            conv.archivada = False
            update_fields.append('archivada')
        if update_fields:
            conv.save(update_fields=update_fields)

        from django.urls import reverse
        return redirect(f"{reverse('whatsapp:inbox')}?conv={conv.pk}")


class AsignarAgenteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not request.user.can_see_all:
            return JsonResponse({'ok': False, 'error': 'Sin permisos'}, status=403)
        conv = get_object_or_404(Conversacion, pk=pk)
        agente_id = request.POST.get('agente_id') or None
        conv.agente_id = agente_id
        conv.save(update_fields=['agente_id'])
        agente_nombre = ''
        if agente_id:
            try:
                u = User.objects.get(pk=agente_id)
                agente_nombre = u.get_full_name() or u.username
            except User.DoesNotExist:
                pass
        return JsonResponse({'ok': True, 'agente_nombre': agente_nombre})


class ArchivarConversacionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        conv = get_object_or_404(_get_convs_qs(request.user), pk=pk)
        conv.archivada = True
        conv.save(update_fields=['archivada'])
        return JsonResponse({'ok': True})


class DesarchivarConversacionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        conv = get_object_or_404(Conversacion, pk=pk)
        conv.archivada = False
        conv.save(update_fields=['archivada'])
        return JsonResponse({'ok': True})


class MarcarLeidoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        Conversacion.objects.filter(pk=pk).update(mensajes_no_leidos=0)
        return JsonResponse({'ok': True})


class BorrarMensajeView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from .models import Mensaje
        msg = get_object_or_404(Mensaje, pk=pk)
        conv = msg.conversacion
        if not request.user.can_see_all and conv.agente != request.user:
            return JsonResponse({'ok': False, 'error': 'Sin permisos'}, status=403)
        if msg.direccion != Mensaje.DIR_SALIENTE:
            return JsonResponse({'ok': False, 'error': 'Solo mensajes enviados'}, status=400)
        if msg.borrado:
            return JsonResponse({'ok': True})
        from .sender import delete_message
        if msg.whatsapp_message_id:
            delete_message(msg.whatsapp_message_id, conv.telefono)
        Mensaje.objects.filter(pk=pk).update(borrado=True, contenido='')
        return JsonResponse({'ok': True})


class AbrirConversacionView(LoginRequiredMixin, View):
    """Cuando el agente abre una conv pendiente, la pasa a 'abierta'."""
    def post(self, request, pk):
        Conversacion.objects.filter(pk=pk, estado=Conversacion.ESTADO_PENDIENTE).update(
            estado=Conversacion.ESTADO_ABIERTA
        )
        return JsonResponse({'ok': True})


class EnviarMediaView(LoginRequiredMixin, View):
    MAX_SIZE_MB = 16

    def post(self, request, pk):
        import base64
        from .sender import send_media_message, get_mediatype

        conv = get_object_or_404(Conversacion, pk=pk)
        archivo = request.FILES.get('archivo')
        caption = request.POST.get('caption', '').strip()
        is_ptt = request.POST.get('ptt') == '1'

        if not archivo:
            return JsonResponse({'ok': False, 'error': 'No se recibió ningún archivo.'}, status=400)

        max_bytes = self.MAX_SIZE_MB * 1024 * 1024
        if archivo.size > max_bytes:
            return JsonResponse({'ok': False, 'error': f'El archivo supera los {self.MAX_SIZE_MB}MB.'}, status=400)

        import os
        from django.conf import settings as dj_settings

        mime = archivo.content_type or 'application/octet-stream'
        mediatype = get_mediatype(mime)
        filename = archivo.name or 'archivo'

        # Guardar localmente para tener URL de referencia
        local_url = ''
        try:
            upload_dir = os.path.join(dj_settings.MEDIA_ROOT, 'uploads', f'conv_{pk}')
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = re.sub(r'[^\w.\-]', '_', filename)
            local_path = os.path.join(upload_dir, safe_name)
            archivo.seek(0)
            with open(local_path, 'wb') as f:
                for chunk in archivo.chunks():
                    f.write(chunk)
            local_url = f'{dj_settings.MEDIA_URL}uploads/conv_{pk}/{safe_name}'
            archivo.seek(0)
        except Exception as save_err:
            logger.warning('No se pudo guardar archivo localmente: %s', save_err)

        try:
            raw = archivo.read()
            b64 = base64.b64encode(raw).decode('utf-8')
            # Evolution API v2: some versions need raw base64, others need data URI
            # Try raw base64 first; if falla, reintentar con data URI
            send_fns = []
            if is_ptt and mediatype == 'audio':
                from .sender import send_whatsapp_audio
                send_fns.append(lambda data: send_whatsapp_audio(conv.telefono, data, filename=filename))
            send_fns.append(lambda data: send_media_message(
                conv.telefono, data, mediatype, filename=filename, caption=caption,
            ))

            result = None
            last_exc = None
            for send_fn in send_fns:
                for media_data in [b64, f'data:{mime};base64,{b64}']:
                    try:
                        result = send_fn(media_data)
                        break
                    except Exception as exc:
                        last_exc = exc
                        logger.warning('Retry con formato alternativo para media...')
                if result is not None:
                    break
            if result is None:
                raise last_exc
            msg_id = result.get('id', '')
        except Exception as e:
            logger.error('Error enviando media a %s: %s', conv.telefono, e)
            return JsonResponse({'ok': False, 'error': f'Error al enviar el archivo: {str(e)[:100]}'}, status=500)

        # Registrar en DB
        tipo_map = {'image': Mensaje.TIPO_IMAGEN, 'video': Mensaje.TIPO_VIDEO,
                    'audio': Mensaje.TIPO_AUDIO, 'document': Mensaje.TIPO_DOCUMENTO}
        msg = Mensaje.objects.create(
            conversacion=conv,
            whatsapp_message_id=msg_id,
            direccion=Mensaje.DIR_SALIENTE,
            tipo=tipo_map.get(mediatype, Mensaje.TIPO_DOCUMENTO),
            contenido=caption,
            media_url=local_url,
            media_mime=mime,
            media_filename=filename,
            status=Mensaje.STATUS_ENVIADO,
            timestamp=timezone.now(),
            enviado_por=request.user,
        )
        Conversacion.objects.filter(pk=conv.pk).update(ultimo_mensaje_at=timezone.now())

        return JsonResponse({
            'ok': True,
            'mensaje': {
                'pk': msg.pk,
                'tipo': msg.tipo,
                'contenido': caption,
                'media_filename': filename,
                'media_mime': mime,
                'timestamp': msg.timestamp.strftime('%H:%M'),
                'enviado_por': request.user.get_full_name() or request.user.username,
            }
        })


class BotToggleView(LoginRequiredMixin, View):
    def post(self, request, pk):
        conv = get_object_or_404(Conversacion, pk=pk)
        bot_type = request.POST.get('bot_type', '')
        activo = request.POST.get('activo', 'true').lower() == 'true'
        if bot_type == 'crm':
            conv.bot_crm_activo = activo
            conv.save(update_fields=['bot_crm_activo'])
        elif bot_type == 'n8n':
            conv.bot_n8n_activo = activo
            conv.save(update_fields=['bot_n8n_activo'])
            if activo:
                from .tasks import liberar_asesor_n8n_task
                liberar_asesor_n8n_task.delay(conv.telefono)
        else:
            return JsonResponse({'ok': False, 'error': 'bot_type inválido'}, status=400)
        return JsonResponse({'ok': True, 'bot_type': bot_type, 'activo': activo})


class PlantillaListView(LoginRequiredMixin, ListView):
    model = PlantillaHSM
    template_name = 'whatsapp/plantilla_list.html'
    context_object_name = 'plantillas'
    paginate_by = 25


class PlantillaCreateView(LoginRequiredMixin, View):
    template_name = 'whatsapp/plantilla_form.html'

    def get(self, request):
        return render(request, self.template_name, {'data': {'nombre': '', 'cuerpo': ''}})

    def post(self, request):
        nombre = request.POST.get('nombre', '').strip()
        cuerpo = request.POST.get('cuerpo', '').strip()
        if not nombre or not cuerpo:
            messages.error(request, 'Nombre y cuerpo son requeridos.')
            return render(request, self.template_name, {'data': request.POST.dict()})
        vars_raw = request.POST.get('variables_raw', '').strip()
        variables = [v.strip() for v in vars_raw.splitlines() if v.strip()] if vars_raw else []
        PlantillaHSM.objects.create(nombre=nombre, cuerpo=cuerpo, variables=variables)
        messages.success(request, 'Plantilla creada.')
        return redirect('whatsapp:plantilla_list')


class PlantillaUpdateView(LoginRequiredMixin, View):
    template_name = 'whatsapp/plantilla_form.html'

    def get(self, request, pk):
        p = get_object_or_404(PlantillaHSM, pk=pk)
        data = {'nombre': p.nombre or '', 'cuerpo': p.cuerpo or ''}
        return render(request, self.template_name, {'obj': p, 'data': data})

    def post(self, request, pk):
        p = get_object_or_404(PlantillaHSM, pk=pk)
        p.nombre = request.POST.get('nombre', p.nombre).strip()
        p.cuerpo = request.POST.get('cuerpo', p.cuerpo).strip()
        vars_raw = request.POST.get('variables_raw', '').strip()
        p.variables = [v.strip() for v in vars_raw.splitlines() if v.strip()] if vars_raw else []
        p.activa = request.POST.get('activa') == 'on'
        p.save()
        messages.success(request, 'Plantilla actualizada.')
        return redirect('whatsapp:plantilla_list')


class PlantillaDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        get_object_or_404(PlantillaHSM, pk=pk).delete()
        messages.success(request, 'Plantilla eliminada.')
        return redirect('whatsapp:plantilla_list')


class ConfigView(SupervisorRequiredMixin, View):
    template_name = 'whatsapp/config.html'

    def get(self, request):
        try:
            config = ConfiguracionWhatsApp.objects.get(pk=1)
        except ConfiguracionWhatsApp.DoesNotExist:
            config = ConfiguracionWhatsApp()
        from .sender import get_connection_state
        try:
            state = get_connection_state()
        except Exception:
            state = 'error'
        return render(request, self.template_name, {'config': config, 'connection_state': state})

    def post(self, request):
        try:
            config = ConfiguracionWhatsApp.objects.get(pk=1)
        except ConfiguracionWhatsApp.DoesNotExist:
            config = ConfiguracionWhatsApp()
        config.evolution_api_url = request.POST.get('evolution_api_url', '').strip()
        config.evolution_api_key = request.POST.get('evolution_api_key', '').strip()
        config.evolution_instance_name = request.POST.get('evolution_instance_name', '').strip() or 'waply'
        config.webhook_token = request.POST.get('webhook_token', '').strip()
        config.save()
        from .sender import setup_instance_webhook, ensure_instance_exists
        from django.urls import reverse
        webhook_url = request.build_absolute_uri(reverse('whatsapp:webhook'))
        ensure_instance_exists()
        setup_instance_webhook(webhook_url)
        messages.success(request, 'Configuración guardada. Webhook registrado.')
        return redirect('whatsapp:config')


class QRCodeView(SupervisorRequiredMixin, View):
    def get(self, request):
        from .sender import get_connection_state, ensure_instance_exists, trigger_connect
        from django.core.cache import cache
        try:
            ensure_instance_exists()
            state = get_connection_state()
            if state == 'open':
                cache.delete('whatsapp_qr_code')
                return JsonResponse({'connected': True, 'qr_base64': None})
            # Only trigger connect if no QR in cache yet
            qr_text = cache.get('whatsapp_qr_text')
            qr_b64 = cache.get('whatsapp_qr_code')
            if not qr_text:
                trigger_connect()
            return JsonResponse({'connected': False, 'qr_base64': qr_b64, 'qr_code': qr_text})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class ConnectionStatusView(LoginRequiredMixin, View):
    def get(self, request):
        from .sender import get_connection_state
        try:
            state = get_connection_state()
            return JsonResponse({'state': state, 'connected': state == 'open'})
        except Exception as e:
            return JsonResponse({'state': 'error', 'connected': False, 'detail': str(e)})


class LogoutInstanceView(SupervisorRequiredMixin, View):
    def post(self, request):
        action = request.POST.get('action', 'logout')
        from .sender import logout_instance, reset_instance
        try:
            if action == 'reset':
                reset_instance()
            else:
                logout_instance()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class APIEnviarMensajeView(View):
    def post(self, request):
        from django.conf import settings as dj
        from .sender import send_text_message, send_media_message
        api_key = getattr(dj, 'CRM_API_KEY', '')
        if not api_key or request.headers.get('X-Api-Key', '') != api_key:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
        # Aceptar 'phone' o 'telefono' indistintamente
        phone = (data.get('phone') or data.get('telefono') or '').strip()
        message = data.get('message', '').strip()
        media_url = data.get('media_url', '').strip()
        media_type = data.get('media_type', 'image')
        if not phone:
            return JsonResponse({'ok': False, 'error': '"phone" requerido'}, status=400)
        if not phone.startswith('+'):
            phone = '+' + phone
        try:
            if media_url:
                result = send_media_message(phone, media_url, media_type, caption=message)
                tipo = Mensaje.TIPO_IMAGEN
            else:
                result = send_text_message(phone, message)
                tipo = Mensaje.TIPO_TEXTO
            conv, _ = Conversacion.objects.get_or_create(
                telefono=phone,
                defaults={'nombre_contacto': phone, 'origen_conversacion': Conversacion.ORIGEN_SALIENTE},
            )
            msg = Mensaje.objects.create(
                conversacion=conv, direccion=Mensaje.DIR_SALIENTE, tipo=tipo,
                contenido=message, media_url=media_url,
                whatsapp_message_id=result.get('id', ''),
                status=Mensaje.STATUS_ENVIADO, timestamp=timezone.now(),
            )
            conv.ultimo_mensaje_at = timezone.now()
            conv.save(update_fields=['ultimo_mensaje_at'])
            return JsonResponse({'ok': True, 'message_id': result.get('id', ''), 'conversacion_id': conv.pk})
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class APIContactoView(View):
    """
    Crea o actualiza un contacto en el CRM con datos adicionales.
    Los campos extra se guardan como CampoPersonalizado + ValorCampo.
    POST /whatsapp/api/contacto/
    Body: {
        "phone": "+549...",
        "nombre": "Juan García",        (opcional)
        "email": "juan@mail.com",       (opcional)
        "notas": "...",                 (opcional)
        "campos": {                     (opcional)
            "localidad": "Canning",
            "origen": "whatsapp",
            "obra_social": "UP"
        }
    }
    """
    def post(self, request):
        from django.conf import settings as dj
        api_key = getattr(dj, 'CRM_API_KEY', '')
        if not api_key or request.headers.get('X-Api-Key', '') != api_key:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

        phone = (data.get('phone') or data.get('telefono') or '').strip()
        if not phone:
            return JsonResponse({'ok': False, 'error': '"phone" requerido'}, status=400)
        if not phone.startswith('+'):
            phone = '+' + phone

        from apps.contacts.models import Contacto, CampoPersonalizado, ValorCampo
        import re as _re

        nombre = data.get('nombre') or data.get('nombre_completo') or ''
        email = data.get('email', '')
        notas = data.get('notas', '')

        defaults = {}
        if nombre:
            defaults['nombre'] = nombre
        if email:
            defaults['email'] = email
        if notas:
            defaults['notas'] = notas

        contacto, created = Contacto.objects.get_or_create(
            telefono=phone,
            defaults={**defaults, 'nombre': nombre or phone},
        )
        if not created and defaults:
            for k, v in defaults.items():
                if v:
                    setattr(contacto, k, v)
            contacto.save()

        # Campos personalizados
        campos_data = data.get('campos') or {}
        campos_guardados = []
        for slug, valor in campos_data.items():
            if not slug or valor is None:
                continue
            slug_clean = _re.sub(r'[^\w]', '_', slug.lower().strip())[:100]
            etiqueta = slug.replace('_', ' ').title()
            campo, _ = CampoPersonalizado.objects.get_or_create(
                nombre=slug_clean,
                defaults={'etiqueta': etiqueta, 'tipo': 'text'},
            )
            ValorCampo.objects.update_or_create(
                contacto=contacto, campo=campo,
                defaults={'valor': str(valor)},
            )
            campos_guardados.append(slug_clean)

        # Vincular conversación si existe y obtener agente asignado
        agente_id = None
        agente_username = None
        agente_nombre = None
        try:
            conv = Conversacion.objects.filter(telefono=phone, contacto__isnull=True).first()
            if conv:
                conv.contacto = contacto
                if nombre and not conv.nombre_contacto:
                    conv.nombre_contacto = nombre
                conv.save(update_fields=['contacto', 'nombre_contacto'])

            # Buscar agente en la conversación activa (ya vinculada o recién vinculada)
            conv_activa = Conversacion.objects.filter(
                telefono=phone, archivada=False
            ).select_related('agente').order_by('-ultimo_mensaje_at').first()
            if conv_activa and conv_activa.agente:
                ag = conv_activa.agente
                agente_id = ag.pk
                agente_username = ag.username
                agente_nombre = f"{ag.first_name} {ag.last_name}".strip() or ag.username
        except Exception:
            pass

        return JsonResponse({
            'ok': True,
            'contacto_id': contacto.pk,
            'created': created,
            'campos_guardados': campos_guardados,
            'agente_id': agente_id,
            'agente_username': agente_username,
            'agente_nombre': agente_nombre,
        })


@method_decorator(csrf_exempt, name='dispatch')
class APIArchivarConversacionView(View):
    """
    Archiva (o desarchiva) una conversación desde el bot / n8n.
    POST /whatsapp/api/archivar/
    Body: {"phone": "+549..."} o {"conversation_id": 42}
          Opcional: {"archivar": false} para desarchivar (default: true)
    Header: X-Api-Key: <CRM_API_KEY>
    """
    def post(self, request):
        from django.conf import settings as dj
        api_key = getattr(dj, 'CRM_API_KEY', '')
        if not api_key or request.headers.get('X-Api-Key', '') != api_key:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

        conv = None
        if data.get('conversation_id'):
            conv = Conversacion.objects.filter(pk=data['conversation_id']).first()
        elif data.get('phone'):
            phone = data['phone']
            if not phone.startswith('+'): phone = '+' + phone
            conv = Conversacion.objects.filter(telefono=phone).first()

        if not conv:
            return JsonResponse({'ok': False, 'error': 'Conversación no encontrada'}, status=404)

        archivar = data.get('archivar', True)
        Conversacion.objects.filter(pk=conv.pk).update(archivada=archivar)
        return JsonResponse({'ok': True, 'conversation_id': conv.pk, 'archivada': archivar})


@method_decorator(csrf_exempt, name='dispatch')
class APIBotToggleExternoView(View):
    """
    n8n puede prender/apagar el bot via API sin sesión de usuario.
    POST /whatsapp/api/bot/
    Body: {"conversation_id": 42, "activo": false}
      o:  {"phone": "+549...", "activo": true}
    """
    def post(self, request):
        from django.conf import settings as dj
        api_key = getattr(dj, 'CRM_API_KEY', '')
        if not api_key or request.headers.get('X-Api-Key', '') != api_key:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

        activo = data.get('activo', False)
        conv = None
        if data.get('conversation_id'):
            conv = Conversacion.objects.filter(pk=data['conversation_id']).first()
        elif data.get('phone'):
            phone = data['phone']
            if not phone.startswith('+'): phone = '+' + phone
            conv = Conversacion.objects.filter(telefono=phone).first()

        if not conv:
            return JsonResponse({'ok': False, 'error': 'Conversación no encontrada'}, status=404)

        Conversacion.objects.filter(pk=conv.pk).update(bot_n8n_activo=activo)
        return JsonResponse({'ok': True, 'conversation_id': conv.pk, 'bot_n8n_activo': activo})


@method_decorator(csrf_exempt, name='dispatch')
class APIHandoffView(View):
    """
    n8n llama a este endpoint cuando el bot termina y quiere pasar la conv a un agente.
    Body JSON: {"conversation_id": 123}  o  {"phone": "+549..."}
    Header: X-Api-Key: <CRM_API_KEY>
    """
    def post(self, request):
        from django.conf import settings as dj
        api_key = getattr(dj, 'CRM_API_KEY', '')
        if not api_key or request.headers.get('X-Api-Key', '') != api_key:
            return JsonResponse({'ok': False, 'error': 'Unauthorized'}, status=401)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

        conv = None
        conv_id = data.get('conversation_id')
        phone = data.get('phone', '').strip()

        if conv_id:
            conv = Conversacion.objects.filter(pk=conv_id).first()
        elif phone:
            if not phone.startswith('+'):
                phone = '+' + phone
            conv = Conversacion.objects.filter(telefono=phone).first()

        if not conv:
            return JsonResponse({'ok': False, 'error': 'Conversación no encontrada'}, status=404)

        # Marcar como pendiente de agente y desactivar bot
        Conversacion.objects.filter(pk=conv.pk).update(
            estado=Conversacion.ESTADO_PENDIENTE,
            bot_n8n_activo=False,
        )
        logger.info('Handoff bot→agente para conv %s', conv.pk)
        return JsonResponse({'ok': True, 'conversation_id': conv.pk, 'estado': 'pendiente'})
